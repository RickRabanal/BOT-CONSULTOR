import inspect, urllib3, paramiko, sys, io, random, xlwt, pyodbc, os, shutil, tkinter, pyperclip, time, openpyxl, webbrowser, smtplib, ssl, subprocess, csv, xlrd, math, PIL, pytesseract, comtypes.client, threading, time, pyperclip, sys, openpyxl, tkinter, mysql.connector, pyodbc
from win32 import win32gui, win32process, win32api
from win32.lib import win32con
from tkinter.filedialog import askopenfilename
import pyscreenshot as img
import cv2 as cv
from tkinter import ttk, messagebox
from openpyxl.styles import colors
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.cell import Cell
from datetime import date, datetime, timedelta
from datetime import timedelta
from fpdf import FPDF
from xlutils.copy import copy
from win32com.client import Dispatch
from PyPDF2 import PdfFileReader, PdfFileWriter, PdfFileMerger
import fitz, warnings
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from email.mime.multipart import MIMEMultipart#
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email.utils import formatdate
from email import encoders
from PIL import Image
from time import sleep
import numpy as np
import win32com
from win32com.client import DispatchEx,Dispatch
from exchangelib.protocol import BaseProtocol, NoVerifyHTTPAdapter
import exchangelib, requests
from collections import Counter as Contador
import locale,stat,decimal
from libreria import *
from datetime import datetime as dt, timedelta as td
def Cerrar(): root.destroy()

class ModalWindow(object):
    global but
    def __init__(self, root, Imagen):
        self.entry_id = tkinter.StringVar()
        self.modalWindow = tkinter.Toplevel(root)
        self.modalWindow.title('Captcha')
        self.modalWindow.resizable(False, False)
        self.modalWindow.wm_attributes("-topmost", True)
        self.modalWindow.geometry('330x100+'+str(int((win32api.GetSystemMetrics(0)-420)/2))+'+'+str(int((win32api.GetSystemMetrics(1)-140)/2)))
        self.modalWindow.overrideredirect(1)
        Imagg = tkinter.PhotoImage(file=os.getcwd() + "\\" + Imagen)
        label = tkinter.ttk.Label(self.modalWindow, image=Imagg)
        label.image = Imagg
        label.place(x=20, y=10)
        labeled = tkinter.ttk.Label(self.modalWindow, text='Ingrese Captcha')
        labeled.place(x=180, y=10)
        but = tkinter.ttk.Entry(self.modalWindow, textvariable=self.entry_id)
        but.place(x=180, y=30)
        Aceptar = tkinter.ttk.Button(self.modalWindow, text='Aceptar', command=self.Adios)
        Aceptar.place(x=180, y=60)
        self.modalWindow.rowconfigure(1, minsize = 500)

    def Adios(self):
        self.modalWindow.destroy()

    def GetCaptcha(self):
        text = self.entry_id.get()
        return str(text)

def colnum_string(n):
    string = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        string = chr(65 + remainder) + string
    return string

def Void(folder):
    for root, dirs, files in os.walk(folder):
        try:
            for f in files: os.unlink(os.path.join(root, f));
            for d in dirs: shutil.rmtree(os.path.join(root, d));
        #except: print('Error:', folder)
        except Exception as inst:
            print("Error al eliminar Carpeta o Archivo: ",str(inst))

#Magia    
def DiccionarioSQL(Select, Modo=2, OnHeader=False):
    if Modo==1: #MySQL
        cnxn = mysql.connector.connect(host='172.30.249.12',database='websoluciones',user='desasoluciones',password='desarrollo123#%')
        cursor = cnxn.cursor(buffered=True)
    elif Modo==2: #SQLServer Negocios
        cnxn = pyodbc.connect('DRIVER={SQL Server};SERVER=172.30.251.85;DATABASE=web_visor;UID=rbt_planificacion;PWD=#rbt_Pl4n1f1c!', autocommit=True)
        #cnxn = pyodbc.connect('DRIVER={SQL Server};SERVER=.;DATABASE=web_visor;UID=sa;PWD=Jhon_123', autocommit=True)         
        cursor = cnxn.cursor()
    elif Modo==3: #MySQL Root
        cnxn = mysql.connector.connect(host='172.30.249.11',database='reclamos',user='root',password='oracle123')
        cursor = cnxn.cursor(buffered=True)
    elif Modo==4: #SQLServer Negocios CORR
        cnxn = pyodbc.connect('DRIVER={SQL Server};SERVER=172.30.251.85;DATABASE=corr;UID=rbt_planificacion;PWD=#rbt_Pl4n1f1c!', autocommit=True)
        #cnxn = pyodbc.connect('DRIVER={SQL Server};SERVER=.;DATABASE=corr;UID=sa;PWD=Jhon_123', autocommit=True)         
        cursor = cnxn.cursor()
    cursor.execute(Select)
    if Modo==1 or Modo==3: cnxn.commit()
    elif Modo==2 or Modo==4: cursor.commit()
    Header, Respuesta = [column[0] for column in cursor.description], []
    Resultado = [[x for x in Rows] for Rows in list(cursor.fetchall())]
    Resultado = [a if str(a)!='None' and a!=None and len(str(a))!=0 else '' for a in Resultado] if len(Resultado)>0 else []
    for idx, x in enumerate(Resultado):
        Lista = {y: x[idy] if x[idy]!=None else '' for idy, y in enumerate(Header)}
        Respuesta.append(Lista)
    if OnHeader: return [Respuesta, Header]
    else: return Respuesta

def Execute(Update, Modo=2):
#print(Update)
    if Modo==1: #MySQL
        cnxn = mysql.connector.connect(host='172.30.249.12',database='websoluciones',user='desasoluciones',password='desarrollo123#%')
        cursor = cnxn.cursor(buffered=True)
    elif Modo==2: #SQLServer Negocios
        cnxn = pyodbc.connect('DRIVER={SQL Server};SERVER=172.30.251.85;DATABASE=web_visor;UID=rbt_planificacion;PWD=#rbt_Pl4n1f1c!', autocommit=True)
        #cnxn = pyodbc.connect('DRIVER={SQL Server};SERVER=.;DATABASE=web_visor;UID=sa;PWD=Jhon_123', autocommit=True)        
        cursor = cnxn.cursor()
    elif Modo==3: #MySQL Root
        cnxn = mysql.connector.connect(host='172.30.249.11',database='reclamos',user='root',password='oracle123')
        cursor = cnxn.cursor(buffered=True)
    elif Modo==4: #SQLServer Negocios CORR
        cnxn = pyodbc.connect('DRIVER={SQL Server};SERVER=172.30.251.85;DATABASE=corr;UID=rbt_planificacion;PWD=#rbt_Pl4n1f1c!', autocommit=True)
        #cnxn = pyodbc.connect('DRIVER={SQL Server};SERVER=.;DATABASE=corr;UID=sa;PWD=Jhon_123', autocommit=True)         
    cursor = cnxn.cursor()
    cursor.execute(Update)
    if Modo==1 or Modo==3: cnxn.commit()
    elif Modo==2 or Modo==4: cursor.commit()
    
def DiccionarioExcelXLSX(Excel, Filass=1):
    Writer = openpyxl.load_workbook(Excel, data_only=True)
    Sheet = Writer.active
    Header, Listado = list(n.value for n in Sheet[Filass]), []
    for idx in range(Sheet.max_row):
        Lista = {}
        for idy, y in enumerate(Header):
            if y == None: continue
            value = Sheet.cell(row=idx+1+Filass, column=idy+1).value
            value = value if value != None and value != '' else ''
            value = value if not isinstance(value, datetime) else value.strftime("%d/%m/%Y")
            if len(str(value).replace(" ", ""))==10 and str(value).count('-')==2 and str(value).replace(" ", "").replace("/", "").replace("-", "").isnumeric():
                print("value",value)
                value = str(value).replace(" ", "").split("-")
                value ="/".join(value[::-1])# value[2] + "/" + value[1] + "/" + value[0]
            if len(str(value))==10 and "/" in str(value): value = value
            if len(str(value))<25 and "/" in str(value) and " " in str(value): value = value.split(" ")[0]
            if 'TELEFONO'==y: value = value.replace("T",'')
            Lista[y.upper()] = str(value).replace("'", "")
        Listado.append(Lista)
    Buscar_Columna(Listado[0])# guardar nombre de columna en un diccionario
    return Listado

def DiccionarioExcelXLS(Excel):
    Writer = xlrd.open_workbook(Excel, on_demand=True)
    Sheet = Writer.sheet_by_name(Writer.sheet_names()[0])
    Header, Listado = [y for y in Sheet.row_values(0)], []

    for idx in range(Sheet.nrows):
        Lista = {}
        if Sheet.nrows==idx+1: break
        for idy, y in enumerate(Header):
            value = Sheet.cell_value(rowx=idx+1, colx=idy)
            value = value if value != None and value != '' else ''
            if 'FECHA' in y:
                if len(str(value).replace(" ", ""))==10 and str(value).count('-')==2 and str(value).replace(" ", "").replace("/", "").replace("-", "").isnumeric():
                    value = str(value).replace(" ", "").split("-")
                    value ="/".join(value[::-1])#
                elif len(str(value))==10 and "/" in str(value) and " " not in str(value): value = str(value)
                elif len(str(value))==10 and "/" in str(value) and " " in str(value): value = str(value.split(" ")[0])
                elif len(str(value))<30 and "/" in str(value) and " " in str(value): value = str(value)
                else:
                    py_date = xlrd.xldate.xldate_as_datetime(value, Writer.datemode)
                    value = value if not isinstance(py_date, datetime) else py_date.strftime("%d/%m/%Y")
            elif 'TELEFONO'==y: value = value.replace("T",'')
            Lista[y.upper()] = str(str(value).replace("'",""))
        Listado.append(Lista)
    Buscar_Columna(Listado[0])# guardar nombre de columna en un diccionario
    return Listado

def ResumenPrevio(Lista, Nombre):    
    #El color verde por Andreita
    Listado, Header = [a for a in Lista], [str(a) for a in Lista[0].keys()]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BASE"
    #Columnas
    Red = PatternFill(start_color='00009E72', end_color='00009E72', fill_type='solid')
    Other = PatternFill(start_color='00FFC000', end_color='00FFC000', fill_type='solid')
    Turqueza = PatternFill(start_color='0040C2CF', end_color='0040C2CF', fill_type='solid')
    for idx, x in enumerate(Header):
        ws.cell(row=1, column=idx+1).value = x.upper()
        ws.cell(row=1, column=idx+1).fill = Other
        ws.cell(row=1, column=idx+1).font = Font(bold=True, color=colors.WHITE)
        ws.column_dimensions[colnum_string(idx+1)].width = 17
    #Cuerpo
    for idx, x in enumerate(Listado):
        for idy, y in enumerate(Header): ws.cell(row=idx+2, column=idy+1).value = x[y]            
    Direccion = os.getcwd() + "\\{0}.xlsx".format(Nombre)
    try:
        wb.save(filename=Direccion)
    except Exception as inst:
        print(str(inst))
    return Direccion

def ResumenPrevio_2(Lista, Nombre):
    Reader = xlrd.open_workbook('Formato2.xls', formatting_info=True)
    RSheet = Reader.sheet_by_index(0)
    Cabeceras = [y for y in RSheet.row_values(0)]
    rb = copy(Reader)
    sheetd, style = rb.get_sheet(0),xlwt.XFStyle()
    for idx, x in enumerate(Lista):
        sheetd.write(idx + 1, Cabeceras.index('NUMERO_RESOLUCION'), x['NUMERO_RESOLUCION'])        
        sheetd.write(idx + 1, Cabeceras.index('FECHA_RESOLUCION'), x['FECHA_RESOLUCION'])
        sheetd.write(idx + 1, Cabeceras.index('COD_RECLAMO'), x['COD_RECLAMO'])
        sheetd.write(idx + 1, Cabeceras.index('SERVICIO'), x['SERVICIO'])
        sheetd.write(idx + 1, Cabeceras.index('NOMBRE'), x['NOMBRE'])
        sheetd.write(idx + 1, Cabeceras.index('ANALISTA'), x['ANALISTA'])
        sheetd.write(idx + 1, Cabeceras.index('FECHA_DESPACHO'), x['FECHA_DESPACHO'])
        sheetd.write(idx + 1, Cabeceras.index('DIRECCION'), x['DIRECCION'])        
        sheetd.write(idx + 1, Cabeceras.index('DISTRITO'), x['DISTRITO'])        
        sheetd.write(idx + 1, Cabeceras.index('PROVINCIA'), x['PROVINCIA'])
        sheetd.write(idx + 1, Cabeceras.index('DEPARTAMENTO'), x['DEPARTAMENTO'])
        sheetd.write(idx + 1, Cabeceras.index('NEGOCIO'), x['NEGOCIO'])
        sheetd.write(idx + 1, Cabeceras.index('RESULTADO'), x['RESULTADO'])
        sheetd.write(idx + 1, Cabeceras.index('CORREO'), x['CORREO'])
        sheetd.write(idx + 1, Cabeceras.index('INSTANCIA'), x['INSTANCIA'])
        sheetd.write(idx + 1, Cabeceras.index('NOMBRE_EXCEL'), x['NOMBRE_EXCEL'])        
        sheetd.write(idx + 1, Cabeceras.index('OBSERVACION'), x['OBSERVACION'])
        sheetd.write(idx + 1, Cabeceras.index('FECHA_RECLAMO'), x['FECHA_RECLAMO'])
        sheetd.write(idx + 1, Cabeceras.index('FECHA_CARGA'), x['FECHA_CARGA'])
        sheetd.write(idx + 1, Cabeceras.index('LLAVE'), x['LLAVE'])         
    Direccion = os.getcwd() + "\\{0}.xls".format(Nombre)
    rb.save(Direccion)
    #print('Excel Creado:', Nombre, len(Lista))
    return Direccion

def ResumenPrevio_3(Lista, Nombre):
    Reader = xlrd.open_workbook('Formato3.xls', formatting_info=True)
    RSheet = Reader.sheet_by_index(0)
    Cabeceras = [y for y in RSheet.row_values(0)]
    rb = copy(Reader)
    sheetd, style = rb.get_sheet(0),xlwt.XFStyle()
    for idx, x in enumerate(Lista):
        sheetd.write(idx + 1, Cabeceras.index('TIPO'), x['TIPO'])        
        sheetd.write(idx + 1, Cabeceras.index('NUMERO_RESOLUCION'), x['NUMERO_RESOLUCION'])
        sheetd.write(idx + 1, Cabeceras.index('FECHA_RESOLUCION'), x['FECHA_RESOLUCION'])
        sheetd.write(idx + 1, Cabeceras.index('COD_RECLAMO'), x['COD_RECLAMO'])
        sheetd.write(idx + 1, Cabeceras.index('SERVICIO'), x['SERVICIO'])
        sheetd.write(idx + 1, Cabeceras.index('NOMBRE'), x['NOMBRE'])
        sheetd.write(idx + 1, Cabeceras.index('ANALISTA'), x['ANALISTA'])
        sheetd.write(idx + 1, Cabeceras.index('FECHA_DESPACHO'), x['FECHA_DESPACHO'])        
        sheetd.write(idx + 1, Cabeceras.index('DIRECCION'), x['DIRECCION'])        
        sheetd.write(idx + 1, Cabeceras.index('DISTRITO'), x['DISTRITO'])
        sheetd.write(idx + 1, Cabeceras.index('PROVINCIA'), x['PROVINCIA'])
        sheetd.write(idx + 1, Cabeceras.index('DEPARTAMENTO'), x['DEPARTAMENTO'])        
        sheetd.write(idx + 1, Cabeceras.index('NEGOCIO'), x['NEGOCIO'])
        sheetd.write(idx + 1, Cabeceras.index('RESULTADO'), x['RESULTADO'])
        sheetd.write(idx + 1, Cabeceras.index('CORREO'), x['CORREO'])
    Direccion = os.getcwd() + "\\{0}.xls".format(Nombre)
    rb.save(Direccion)
    #print('Excel Creado:', Nombre, len(Lista))
    return Direccion


def TablaHTML(Lista):
    def Colores(a):
        if a.upper() == 'NUEVO' or a.upper() == 'ANTIGUO': return '0, 176, 80'
        elif a.upper() == 'DENTRO DE PLAZO' or a.upper() == 'FUERA DE PLAZO': return '255, 192, 0'
        return '0, 176, 240'
    Tablita = '<table width="1200" style="border-collapse:collapse;width:600pt"><colgroup><col width="400" span="{0}" style="mso-width-source:userset;mso-width-alt:7862; width:300pt"></colgroup>'.format(len(Lista))
    Header = '<td height="26" width="300pt" style="padding-top: 1px; padding-right: 1px; padding-left: 1px; color: black; font-size: 11pt; font-family: Calibri, sans-serif; vertical-align: bottom; border: none; color: white; font-weight: 700; font-family: Cambria, serif; text-align: center; vertical-align: middle; border: 0.5pt solid rgb(216, 216, 216); background: rgb({1}); border: 1px solid rgb(212, 212, 212); height:20.1pt; width:161pt">{0}</td>'
    Tablita = Tablita + '<tbody><tr height="26" style="mso-height-source:userset;height:20.1pt">{0}</tr>'.format("".join([Header.format(a, Colores(a)) for a in Lista[0].keys()]))
    Body = '<td height="26" style="padding-top: 1px; padding-right: 1px; padding-left: 1px; color: black; font-size: 11pt; font-family: Calibri, sans-serif; vertical-align: bottom; border: none; color: {2}; font-family: Cambria, serif; text-align: center; vertical-align: middle; border: 0.5pt solid rgb(216, 216, 216); background: {1}; border: 1px solid rgb(212, 212, 212); height:20.1pt; border-top:none">{0}</td>'
    Tablita = Tablita + "".join(['<tr height="26" style="mso-height-source:userset;height:20.1pt">{0}</tr>'.format("".join([Body.format(str(b), ("rgb(0, 176, 240)" if a['GRUPOS']=='TOTAL' else "white") if "GRUPOS" in list(a.keys()) else "white", ("white" if a['GRUPOS']=='TOTAL' else "black") if "GRUPOS" in list(a.keys()) else "black") for b in a.values()])) for a in Lista])
    Tablita = Tablita + "</tbody></table>"
    return Tablita

def Consolidado(Listado):
    Nombre = "{0} {1}.xls".format(Listado[0]['CANAL'], str(date.today().strftime("%d.%m.%Y")))
    Reader = xlrd.open_workbook('Formato.xls', formatting_info=True)
    RSheet = Reader.sheet_by_index(0)
    Cabeceras = [y for y in RSheet.row_values(0)]
    rb = copy(Reader)
    sheetd, style = rb.get_sheet(0),xlwt.XFStyle()
##    print('Entra al llenado-----------------')
##    print(Listado[0])
    for idx, x in enumerate(Listado):
        sheetd.write(idx + 1, Cabeceras.index('Guía'), x['GUIA'])
        sheetd.write(idx + 1, Cabeceras.index('Resolución'), x['NUMERO_RESOLUCION'])
        style.num_format_str = 'DD/MM/YYYY'
        sheetd.write(idx + 1, Cabeceras.index('Fecha Resolución'), x['FECHA_RESOLUCION'], style)
        sheetd.write(idx + 1, Cabeceras.index('Expediente'), x['COD_RECLAMO'])
        sheetd.write(idx + 1, Cabeceras.index('Teléfono'), x['SERVICIO'].replace("T", ""))
        sheetd.write(idx + 1, Cabeceras.index('Nombre Cliente'), x['NOMBRE'])
        sheetd.write(idx + 1, Cabeceras.index('Analista'), x['ANALISTA'])
        style.num_format_str = 'DD/MM/YYYY'
        sheetd.write(idx + 1, Cabeceras.index('Fecha Despacho'), x['FECHA_DESPACHO'], style)
        sheetd.write(idx + 1, Cabeceras.index('Dirección'), x['DIRECCION'])
        sheetd.write(idx + 1, Cabeceras.index('Distrito'), x['DISTRITO'])
        sheetd.write(idx + 1, Cabeceras.index('Provincia'), x['PROVINCIA'])
        sheetd.write(idx + 1, Cabeceras.index('Departamento'), x['DEPARTAMENTO'])
        sheetd.write(idx + 1, Cabeceras.index('Anexo'), x['NEGOCIO'])
        sheetd.write(idx + 1, Cabeceras.index('Tipo Reenvío'), '' if x['TIENE_RECIBO']=='NO' else 'RECIBO')
        sheetd.write(idx + 1, Cabeceras.index('Tipo Solucion'), x['RESULTADO'])
        sheetd.write(idx + 1, Cabeceras.index('Número Reenvío'), '')
        sheetd.write(idx + 1, Cabeceras.index('Correo electrónico'), x['CORREO'])
        style.num_format_str = 'DD/MM/YYYY'
        sheetd.write(idx + 1, Cabeceras.index('Fecha Reclamo'), x['FECHA_RECLAMO'], style)
        sheetd.write(idx + 1, Cabeceras.index('Canal Despacho'), x['CANAL'])
        sheetd.write(idx + 1, Cabeceras.index('Instancia'), x['INSTANCIA'])
        
    rb.save(os.getcwd() + '\\Resultado\\' + Nombre)
    print('Excel Creado:', Nombre, len(Listado))

def GuiaMax():
    GuiaMax = DiccionarioSQL("SELECT MAX(GUIA) AS GUIA FROM robot_despacho WHERE YEAR(TRY_CONVERT(DATE, FECHA_DESPACHO, 103)) = YEAR(GETDATE())", 4)
    if len(GuiaMax)>0: GuiaMax = int(GuiaMax[0]['GUIA'].split('-')[0]) if GuiaMax[0]['GUIA']!='' else 0
    else: GuiaMax = 0
    return GuiaMax


def CreaPDF(x):
    try:
        pdf_writer = PdfFileWriter()
        for page in x['PDF']: pdf_writer.addPage(page)
        Lista = list(filter(lambda f: f.endswith(('.pdf','.PDF')), [f for f in os.listdir(Ruta_descargas)]))
        x['TIENE_RECIBO'] = 'NO'
        if x['COD_RECLAMO'] + ".pdf" in Lista:
            x['TIENE_RECIBO'] = 'SI'
            pdf = PdfFileReader(Ruta_descargas + "\\" + x['COD_RECLAMO'] + ".pdf")
            pdf._override_encryption = True
            pdf._flatten()
            #if pdf.isEncrypted: print('ENCRIPTADO:', x['COD_RECLAMO']); return x;
            for page in range(pdf.getNumPages()):
                pdf_writer.addPage(pdf.getPage(page))
        with open(os.getcwd()+ "\\Resultado\\" + x['CANAL'] + "\\" + x['NOMBRE_PDF'], 'wb') as out: pdf_writer.write(out)
        Execute("UPDATE robot_despacho SET NOMBRE_ARCHIVO='{0}', GUIA='{1}' WHERE ID={2}".format(x['NOMBRE_PDF'], x['GUIA'], x['ID']), 4)
    except Exception as e:
        print(f"ERROR: '{x['COD_RECLAMO']}'"," ,Exception:",e);
        return x;
    return x

def UnirPDFs(Listado):
    Nombre, merger, Canal = "CONSOLIDADO {0}.pdf".format(Listado[0]['CANAL']), PdfFileMerger(), Listado[0]['CANAL']
    Lista = list(filter(lambda f: f.endswith(('.pdf','.PDF')), [f for f in os.listdir(Ruta_descargas)]))
    for x in Listado:
        try:
            if os.path.exists(os.getcwd() + "\\Resultado\\" + x['CANAL'] + "\\" + x['NOMBRE_PDF']):
                fd = open(os.getcwd() + "\\Resultado\\" + x['CANAL'] + "\\" + x['NOMBRE_PDF'], 'rb')
                merger.append(PdfFileReader(fd))
                if x['COD_RECLAMO'] + ".pdf" in Lista:
                    pdf = PdfFileReader(Ruta_descargas + "\\" + x['COD_RECLAMO'] + ".pdf")
                    for page in range(pdf.getNumPages()):
                        merger.append(pdf.getPage(page)) 
        finally:
            if fd: fd.close()
    for x in Listado: os.remove(os.getcwd() + "\\Resultado\\" + x['CANAL'] + "\\" + x['NOMBRE_PDF'])
    merger.write(os.getcwd() + "\\Resultado\\" + str(Canal) + '\\' + Nombre)

def InicioPortal():
    global browser
    usuario, Contraseña = 'JMAUROS', '40780'
    download_dir = os.getcwd()+"\\Descargas"
    chrome_options = Options()
    chrome_options.add_argument("--disable-notifications")
    chrome_options.add_argument('--no-sandbox')
    chrome_options.add_argument('--verbose')
    chrome_options.add_experimental_option("prefs", {
          "download.prompt_for_download": False,
          "download.directory_upgrade": True,
          "safebrowsing.enabled": True,
          "browser.helperApps.alwaysAsk.openFile": False,
          "profile.default_content_setting_values.automatic_downloads": 1,
          "plugins.always_open_pdf_externally": True,
          "plugins.plugins_list": [{"enabled": False, "name": "Chrome PDF Viewer"}],
          "download.default_directory": download_dir, "download.extensions_to_open": "applications/pdf"
    })
    chrome_options.add_argument('--disable-gpu')
    chrome_options.add_argument('--disable-software-rasterizer')
    chrome_options.add_argument("--start-maximized")
    browser = webdriver.Chrome(os.getcwd() + "//chromedriver", options = chrome_options)
    browser.get('http://10.226.4.39:8080/VB_WEB/login.jsp')
    browser.find_element_by_name("userid").send_keys(usuario)
    browser.find_element_by_name("password").send_keys(Contraseña)

    browser.execute_script("window.scrollTo(0, document.body.scrollHeight);")
    Captcha = browser.find_elements_by_xpath('//*[@id="content"]/table/tbody/tr/td/form/div/table/tbody/tr[6]/td[1]/img')[0]
    img_str = Captcha.screenshot_as_png
    img = Image.open(io.BytesIO(img_str))
    img_array = np.asarray(img)

    cv.imwrite('Captcha.png',img_array)
    modal = ModalWindow(root,'Captcha.png')
    root.wait_window(modal.modalWindow)

    browser.find_element_by_name("answer").send_keys(modal.GetCaptcha())
    browser.find_element_by_name("Submit").click()
    t = 0
    while True:
        time.sleep(0.5)
        if len(browser.find_elements_by_xpath('//*[@id="content"]/table/tbody/tr/td/form/div/table/tbody/tr[6]/td[1]/img'))==0: break
        elif t>5: goto .inicio
        t+=1
    if not(os.path.exists(os.getcwd().split('Desktop')[0]+"Dataset")): os.makedirs(os.getcwd().split('Desktop')[0]+"Dataset")
    shutil.move(os.getcwd()+'\\Captcha.png',os.getcwd().split('Desktop')[0]+"Dataset\\{0}.png".format(modal.GetCaptcha()))
    return browser

def DoRecibos(x):
    global tabla,codigo,codigo1,pos,tabla
    browser.get("http://10.226.4.39:8080/VB_WEB/BusquedaTipoDoc.jsp")
    if x['SISTEMA'] == 'FIJA':
        browser.find_element_by_xpath("//*[@id='tipo_producto']/option[text()='{0}']".format('Movistar Fijo')).click()
        browser.find_element_by_name("nroTelef_f").clear()
        browser.find_element_by_name("nroTelef_f").send_keys(str(x['TELEFONO']).zfill(8))
    else:
        browser.find_element_by_xpath("//*[@id='tipo_producto']/option[text()='{0}']".format('Movistar TV')).click()
        browser.find_element_by_name("nroCliente_c").clear()
        browser.find_element_by_name("nroCliente_c").send_keys(str(x['TELEFONO']).zfill(8))
    browser.find_elements_by_xpath('//*[@id="content"]/table/tbody/tr/td/div/form/fieldset/input')[0].click()
    tabla = browser.find_elements_by_xpath('//*[@id="content"]/table/tbody/tr/td/div/form/table')[0]
    tabla = tabla.text.split("\n")
    cabecera, body = tabla[:2], tabla[2:]
    if not('Disculpe pero no existen recibos para los datos ingresados.' in tabla or 'Estimado cliente su recibo se encuentra en proceso' in tabla) and not(os.path.exists(os.getcwd()+"\\Descargas\\{0}.pdf".format(x['COD_RECLAMO']))):
        pos=[idx for idx,x in enumerate(body) if x.count('/')==2]+[None]
        tabla=[body[pos[idx]:pos[idx+1]] for idx in range(len(pos)-1)]
        codigo=[x[1] if x[1].count('-')==1 else None for x in tabla]
        codigo1=[idx  for idx,x in enumerate(codigo) if x[:2] in ('s0','00')]
        if codigo1==[]:
            codigo1=[idx  for idx,x in enumerate(codigo) if x[0]=='L']
            direccion='//*[@id="content"]/table/tbody/tr/td/div/form/table/tbody/tr[{0}]/td[4]'.format(str(codigo1[0]+3))
        else:
            direccion='//*[@id="content"]/table/tbody/tr/td/div/form/table/tbody/tr[{0}]/td[6]'.format(str(codigo1[0]+3))
        tabla=browser.find_elements_by_xpath(direccion)[0]
        pdf=browser.find_element_by_xpath(direccion+"/div/a")
        print(pdf.get_attribute('onclick'))
        url = pdf.get_attribute('href').split("/VB_WEB")[0]+pdf.get_attribute('onclick').split("'")[1]
        browser.get(url)
        if 'Estimado cliente su recibo se encuentra en proceso' in browser.page_source: return
        while True:
            if os.path.exists(os.getcwd()+"\\Descargas\\factura.pdf"): break
            time.sleep(0.3)
        print("factura.pdf:",os.path.exists(os.getcwd()+"\\Descargas\\factura.pdf")) 
        os.rename(os.getcwd()+"\\Descargas\\factura.pdf",os.getcwd()+"\\Descargas\\{0}.pdf".format(x['COD_RECLAMO']))

def Recibos():
    browser = InicioPortal()
    Recibose = ["'" + str(a).replace('.pdf', '') + "'" for a in list(filter(lambda f: f.endswith(('.pdf','.PDF')), [f for f in os.listdir(Ruta_descargas)]))]
    query = """
            SELECT DISTINCT
            CASE WHEN SISTEMA in ('WST','TV') THEN  REPLICATE('0', (8 - LEN(SERVICIO))) + SERVICIO  WHEN SISTEMA IN ('FENIX','WSF') THEN SERVICIO  END  AS TELEFONO,
            COD_RECLAMO,
            CASE WHEN SISTEMA in ('WST','TV') THEN 'TV' WHEN SISTEMA IN ('FENIX','WSF') THEN 'FIJA'  END AS SISTEMA
            FROM robot_Cartas_fija_primera WHERE MOTIVO_RECLAMO='#C88' AND
            (CONVERT(VARCHAR,FECHA_CARGA,103) = CONVERT(VARCHAR,DATEADD(DAY,-2,GETDATE()),103)
            or CONVERT(VARCHAR,FECHA_CARGA,103) = CONVERT(VARCHAR,GETDATE(),103)) AND COD_RECLAMO NOT IN ({0})
    """.format(",".join(Recibose))
    #print (query)
    #Lista = DiccionarioSQL("SELECT DISTINCT SERVICIO AS TELEFONO, COD_RECLAMO, CASE WHEN SISTEMA='WST' THEN 'TV' ELSE 'FIJA' END AS SISTEMA FROM robot_Cartas_fija_primera WHERE MOTIVO_RECLAMO='#C88' AND FECHA_CARGA=CONVERT(VARCHAR,GETDATE(),103) AND COD_RECLAMO NOT IN ({0})".format(",".join(Recibose)))
    Lista = DiccionarioSQL(query)
    pb.configure(maximum=len(Lista))
    progress.set(1)
    print("Inicio!")
    for idx, x in enumerate(Lista):
        progress.set(idx)
        Cuenta.config(text=str(idx+1)+"/"+str(len(Lista)))
        root.wm_attributes("-topmost", True)
        root.update_idletasks()
        root.update()
        DoRecibos(x)
    print('Listo!')
    messagebox.showinfo("Mensaje", "Terminó")

def Word2PDF(f):
    Word= DispatchEx("Word.Application")
    Word.Visible = False
    Word.DisplayAlerts = False
    doc=Word.Documents.Open(os.path.abspath(Ruta_Despacho + "\\" + f))
    doc.SaveAs(os.path.abspath(Ruta_Despacho + "\\test.pdf"),FileFormat=17)
    Word.DisplayAlerts = True
    Word.Quit()
    del Word

##################################################################################################################
##################################################################################################################

def Buscar_Columna(a):
    val=DiccionarioSQL("select Columna_variable,Columna_correcta from sa_diccionario_columnas", 4)
    columna={x['Columna_variable']:x['Columna_correcta'] for x in val}
    dicti={x:[y for y in columna.keys() if columna[y]==x] for x in dict.fromkeys(columna.values()).keys()}    
    for x in dicti.keys():
        for y in a.keys():
            if y in dicti[x]:
                Change[x]= y
                break

##def Buscar_Columna(a):
##    def Buscar_ColumnaB(a,columna):
##        z = columna if columna != 'CANAL' else ''
##        for b in list(a.keys()):
##            for c in  DiccionarioSQL("select  Columna_variable,Columna_correcta from sa_diccionario_columnas where Columna_correcta='"+columna+"'  ", 4):
##                    #print(c,b)
##                if c['Columna_variable']==b and c['Columna_correcta']==columna:
##                    z= c['Columna_variable']
##        return z
##
##    #Change = {"NUMERO_RESOLUCION":"","FECHA_RESOLUCION":"","CODIGO_RECLAMO":"","SERVICIO":"","NOMBRE":"","ANALISTA":"","DIRECCION":"","DISTRITO":"","PROVINCIA":"","DEPARTAMENTO":"","ANEXO_INSCRIPCION":"","FECHA_RECLAMO":"","MONTO_RECLAMO":"","TIPO_ENVIO":"","RESULTADO":"","CORREO":"","CANAL":"","INSTANCIA":"","NEGOCIO":"","CODIGO_RECLAMO_1RA":""}
##    Change['NUMERO_RESOLUCION']=Buscar_ColumnaB(a,'NUMERO_RESOLUCION')
##    Change['FECHA_RESOLUCION']=Buscar_ColumnaB(a,'FECHA_RESOLUCION')
##    Change['CODIGO_RECLAMO']=Buscar_ColumnaB(a,'CODIGO_RECLAMO')
##    Change['ANALISTA']=Buscar_ColumnaB(a,'ANALISTA')
##    Change['DIRECCION']=Buscar_ColumnaB(a,'DIRECCION')
##    Change['DISTRITO']=Buscar_ColumnaB(a,'DISTRITO')
##    Change['PROVINCIA']=Buscar_ColumnaB(a,'PROVINCIA')
##    Change['DEPARTAMENTO']=Buscar_ColumnaB(a,'DEPARTAMENTO')
##    Change['MONTO_RECLAMO']=Buscar_ColumnaB(a,'MONTO_RECLAMO')
##    Change['TIPO_ENVIO']=Buscar_ColumnaB(a,'TIPO_ENVIO')
##    Change['INSTANCIA']=Buscar_ColumnaB(a,'INSTANCIA')
##    Change['NEGOCIO']=Buscar_ColumnaB(a,'NEGOCIO')
##    Change['SERVICIO']=Buscar_ColumnaB(a,'SERVICIO')
##    Change['NOMBRE']=Buscar_ColumnaB(a,'NOMBRE')
##    Change['ANEXO_INSCRIPCION']=Buscar_ColumnaB(a,'ANEXO_INSCRIPCION')
##    Change['FECHA_RECLAMO']=Buscar_ColumnaB(a,'FECHA_RECLAMO')
##    Change['RESULTADO']=Buscar_ColumnaB(a,'RESULTADO')
##    Change['CORREO']=Buscar_ColumnaB(a,'CORREO')    
##    Change['CODIGO_RECLAMO_1RA']=Buscar_ColumnaB(a,'CODIGO_RECLAMO_1RA')
##    Change['CANAL']=Buscar_ColumnaB(a,'CANAL')


def Cambio(a):
    try:
        x = {}
        x['SEGMENTO'] = "RESIDENCIAL"
        x['FECHA_DESPACHO'] = (datetime.today()).strftime("%d/%m/%Y")
        x['NUMERO_RESOLUCION'] = str(a[Change['NUMERO_RESOLUCION']]).replace(' ', '').upper()
        x['FECHA_RESOLUCION'] = a[Change['FECHA_RESOLUCION']]
        x['COD_RECLAMO'] = str(a[Change['CODIGO_RECLAMO']]).replace(' ', '').upper()
        x['ANALISTA'] = a[Change['ANALISTA']].upper()
        x['DIRECCION'] = a[Change['DIRECCION']].upper()
        x['DISTRITO'] = a[Change['DISTRITO']].upper()
        x['PROVINCIA'] = a[Change['PROVINCIA']].upper()
        x['DEPARTAMENTO'] = a[Change['DEPARTAMENTO']].upper()
        x['MONTO_RECLAMO'] = a[Change['MONTO_RECLAMO']]
        x['TIPO_ENVIO'] = a[Change['TIPO_ENVIO']].replace(" ", "").upper()
        x['INSTANCIA'] = a[Change['INSTANCIA']].upper()
        x['NEGOCIO'] = a[Change['NEGOCIO']].upper()
        x['SERVICIO'] = str(a[Change['SERVICIO']]).replace('T', '') 
        x['NOMBRE'] =  a[Change['NOMBRE']].upper()
        x['ANEXO_INSCRIPCION'] = a[Change['ANEXO_INSCRIPCION']]
        
        f_reclamo =  str(a[Change['FECHA_RECLAMO']]) 
        x['FECHA_RECLAMO'] = f_reclamo.split(" ")[0] if ' ' in f_reclamo  else f_reclamo

        Resultado = a[Change['RESULTADO']].replace(" ", "").upper() 
        if "PROCEDENTE" == Resultado: Resultado = "FUNDADO"
        elif "FUNDADO EN PARTE" == Resultado: Resultado = "INFUNDADO"
        elif "IMPROCEDENTE" == Resultado or "INPROCEDENTE" == Resultado: Resultado = "INFUNDADO"
        elif "INFUNDADO" in Resultado: Resultado = "INFUNDADO"
        elif "FUNDADO" in Resultado: Resultado = "FUNDADO"
        x['RESULTADO'] = Resultado

        CorreoValido = a[Change['CORREO']]    
        CorreoValido = CorreoValido !='' and len(CorreoValido.replace(' ', ''))>6 and '@' in CorreoValido and False not in [b not in CorreoValido.lower() for b in ['dominio', 'notiene', 'nohay', 'email', 'email.com', 'sincorreo', '@sms', 'dummy.com', 'sms.com']]
        x['CORREO'] = a[Change['CORREO']].upper() if CorreoValido else '' #Correo

        Reclamo =   a[Change['CODIGO_RECLAMO_1RA']].upper() 
        x['CODIGO_RECLAMO_1RA'] = Reclamo

        if  Change['CANAL']=="":
            canal="FISICO"
        else:    
            canal=a[Change['CANAL']].replace(" ", "").upper()
            canal= 'FISICO' if (canal.upper()=='URBANO' or canal.upper()=='SMS' or canal=='' or canal=='None') else canal 
        x['CANAL'] = canal
        return x
    except Exception as inst:
        print("Error:",inst)
        x = {}
        x['ERROR'] = str(inst)
    return x

def SMTPLib(Subject, Recipients, Body, attachments=None):
    print('inicia')
    class ProxyAdapter(requests.adapters.HTTPAdapter):
        def send(self, *args, **kwargs):
            kwargs['proxies'] = {
                'http': 'telefonica01.gp.inet:8080',
                'https': 'telefonica01.gp.inet:8080',
                'socks': 'telefonica01.gp.inet:1080',
                'ftp': 'telefonica01.gp.inet:8080',
            }
            return super(ProxyAdapter, self).send(*args, **kwargs)
    print('2-------')
    User = DiccionarioSQL("SELECT APLICATIVO, USUARIO, CLAVE FROM robot_accesibilidad WHERE Aplicativo = 'Correo Soluciones'", 4)[0]
    print('3-------')
    BaseProtocol.HTTP_ADAPTER_CLS = ProxyAdapter
    BaseProtocol.HTTP_ADAPTER_CLS = NoVerifyHTTPAdapter
    urllib3.disable_warnings()
    print('4-------')
    ### cambio ###
##    creds = exchangelib.Credentials(User['USUARIO'], User['CLAVE'])
    print('5-------')
    creds = exchangelib.Credentials(User['USUARIO'],'Indica#2022-4')   #User['CLAVE']
    config = exchangelib.Configuration(server='10.226.5.152', credentials=creds, retry_policy=exchangelib.FaultTolerance(max_wait=3600), auth_type=exchangelib.NTLM)
    print('6-------')
    account = exchangelib.Account(primary_smtp_address='solucionesindicadoresperu@movistar.com.pe', autodiscover=False, config = config, access_type=exchangelib.DELEGATE)


    
    print('7-------')
    msj = exchangelib.Message(account=account, subject=Subject, to_recipients=Recipients)
    print('8-------')
    
    msj.body = exchangelib.HTMLBody(Body)
    for attachment_name, attachment_content in attachments or []:
        file = exchangelib.FileAttachment(name=attachment_name, content=attachment_content)
        msj.attach(file)

    msj.send()
    for i in Recipients:
        print('Mensaje Enviado a',i ,':', Subject, " " ,  datetime.now().strftime("%d/%m/%Y %H:%M:%S"))

##    print('Mensaje Enviado:', Subject, " " ,  datetime.now().strftime("%d/%m/%Y %H:%M:%S"))


    










    


def CorreoFaltantes(Faltantes=[]):
    Recipient = [a['FALTANTES'] for a in DiccionarioExcelXLSX('Correos.xlsx') if a['FALTANTES']!='']
    Subject = "Casos no despachados " + time.strftime("%d/%m/%Y")
    fp = open('Mensaje.html')
    if Faltantes!=[]:
        Todo = list(filter(lambda f: f.endswith(('.xls','.XLS','.xlsx','.XLSX')), [f for f in os.listdir(Ruta_Despacho)]))
        Sila = [a['NOMBRE_EXCEL'] for a in DiccionarioSQL("SELECT DISTINCT NOMBRE_EXCEL FROM robot_despacho WHERE FECHA_CARGA=CONVERT(VARCHAR,GETDATE(),103)", 4)]
        Faltantes = [a for a in Todo if a not in Sila]
    Body = fp.read().replace("{0}", "".join(["<p>" + a + "</p>" for a in Faltantes]))
    attachments = []
    with open(ResumenPrevio(DiccionarioSQL(Robot_No_Despachado, 4), 'Casos No Despachados'), 'rb') as f: content = f.read()
    attachments.append(('Casos No Despachados.xlsx', content))
    SMTPLib(Subject, Recipient, Body, attachments)

    
def Correo_No_Despacho():
    print("INICIA")

    

    if len(DiccionarioSQL("SELECT NUMERO_RESOLUCION FROM CORR..robot_despacho WHERE FECHA_CARGA=CONVERT(VARCHAR,GETDATE(),103) AND GUIA IS NULL ", 4))==0:
        messagebox.showinfo("Mensaje", "No hay ningun caso")         
        return
    print("Va a llenar la base")
    Base = """
SELECT NEGOCIO, OBSERVACION,
CASE
WHEN OBSERVACION='DUPLICADO LLAVE' THEN 'Sólo deben enviar un caso y copiar en su carpeta de despacho'
WHEN OBSERVACION='YA DESPACHADO' THEN 'Ninguna acción'
WHEN OBSERVACION='DUPLICADO MISMA NUMERO_RESOLUCION,FECHA_RECLAMO Y DISTINTO (ANALISTA O NOMBRE_EXCEL O RECLAMANTE)' THEN 'Sólo deben enviar un caso y copiar en su carpeta de despacho'
WHEN OBSERVACION='DUPLICADO MISMO NUMERO_RESOLUCION,FECHA_RECLAMO,COD_RECLAMO,RECLAMANTE  DISTINTO ANALISTA' THEN 'Sólo deben enviar un caso y copiar en su carpeta de despacho'
WHEN OBSERVACION='DUP SALIO PRIMERO' THEN 'Ninguna acción' 
WHEN OBSERVACION='NO EXISTE PDF' THEN 'Generar y copiar formato de despacho y pdf en su carpeta de despacho' 
WHEN OBSERVACION='FORMATO' THEN 'Columna Canal solo toma valor Digital o Fisico' 
END ACCION
,CANTIDAD
FROM (
	SELECT NOMBRE_ARCHIVO OBSERVACION,count(NUMERO_RESOLUCION)CANTIDAD,NEGOCIO 
	FROM CORR..robot_despacho
	WHERE FECHA_CARGA=CONVERT(VARCHAR,GETDATE(),103) AND GUIA IS NULL 
	--AND NOMBRE_ARCHIVO<>'YA DESPACHADO'
	GROUP BY NOMBRE_ARCHIVO,NEGOCIO
	)Q
	ORDER BY NEGOCIO
    """
    print("termina base")
    print("rellena select")
    Select = """
	SELECT * FROM (
	SELECT 
	NUMERO_RESOLUCION, FECHA_RESOLUCION, COD_RECLAMO, SERVICIO, NOMBRE, ANALISTA, FECHA_DESPACHO, DIRECCION, DISTRITO, PROVINCIA,
	DEPARTAMENTO, NEGOCIO, RESULTADO, CORREO , INSTANCIA ,NOMBRE_EXCEL,NOMBRE_ARCHIVO OBSERVACION ,FECHA_RECLAMO,FECHA_CARGA,LLAVE
	FROM CORR..robot_despacho
	WHERE FECHA_CARGA=CONVERT(VARCHAR,GETDATE(),103) AND GUIA IS NULL
	UNION ALL 
	SELECT 
	NUMERO_RESOLUCION, FECHA_RESOLUCION, COD_RECLAMO, SERVICIO, NOMBRE, ANALISTA, FECHA_DESPACHO, DIRECCION, DISTRITO, PROVINCIA,
	DEPARTAMENTO, NEGOCIO, RESULTADO, CORREO , INSTANCIA ,NOMBRE_EXCEL,NOMBRE_ARCHIVO OBSERVACION ,FECHA_RECLAMO,FECHA_CARGA,LLAVE
	FROM CORR..robot_despacho
	WHERE LLAVE IN (SELECT LLAVE FROM CORR..robot_despacho	WHERE FECHA_CARGA=CONVERT(VARCHAR,GETDATE(),103) AND GUIA IS NULL AND NOMBRE_ARCHIVO = 'YA DESPACHADO')
	AND FECHA_CARGA <> CONVERT(VARCHAR,GETDATE(),103) 
	)Q ORDER BY LLAVE , OBSERVACION
    """
    print("termina select")
    Lista, fp = DiccionarioSQL(Base, 4), open('Mensaje_no_despachado.html')
    print("lleno lista y fp")
    Subject = "Notificación de Despacho - Casos no despachados " + dt.today().strftime("%d/%m/%Y")
    Recipient = [a['FALTANTES'] for a in DiccionarioExcelXLSX('Correos.xlsx') if a['FALTANTES']!='']
    attachments = []
    with open(ResumenPrevio_2(DiccionarioSQL(Select, 4), 'Casos_no_Despachados'), 'rb') as f: content = f.read()
    attachments.append(('Casos_no_Despachados.xls', content))
    print("entro5")    
    SMTPLib(Subject, Recipient, fp.read().replace("[REEMPLAZAR]", TablaHTML(Lista)), attachments)

def make_excel(resultado,dst,name):# Mejorar funcion
    """
    resultado: resultado del query
    dst: ruta de de destino
    name: nombre del archivo
    """
    #crea un valor de amdocs
##    global rows,r
    print('resultado',len(resultado))
    if len(resultado)==0:return
    if isinstance(resultado[0],dict):
        Header=[x for x in resultado[0]]
        body = [[x[y] for y in x] for x in resultado]
        resultado = [Header] + body
    print(1)
    wb=openpyxl.Workbook()#(write_only=True)
    sheet =wb.active
    sheet.title = "Hoja 1"
    filtro=lambda x: str(int(x))if isinstance(x,decimal.Decimal) else (x.strftime("%d/%m/%Y %H:%M:%S") if isinstance(x,datetime) else ( "" if x==None else x))
    rows = [list(map(filtro,x)) for x in resultado]
    cabecera = openpyxl.styles.PatternFill(start_color='005B9BD5', end_color='005B9BD5', fill_type='solid')
    par=openpyxl.styles.PatternFill(start_color='00DEEBF6', end_color='00DEEBF6', fill_type='solid')
    error=[]
    r=[]
    print(2)
    for row in rows:
        try:
            sheet.append(row)
        except:
            print(row)# error por caracteres especiales
            error.append(row)
    print(3)
    columnas = sheet.max_column
    filas = sheet.max_row
    for y in range(filas,filas+len(error)):
        for x in range(columnas):
            try:
                sheet.cell(row=y+1,column=x+1).value=str(error[y-filas][x]).replace('\x03','').replace('\x13','')
            except:
                r.append(error[y-filas][x])
                print(error[y-filas][x])
    print(4)
    columnas = sheet.max_column
    filas = sheet.max_row
    ## cabeceras color
    for x in range(columnas):
        sheet.cell(row=1, column=x+1).font = openpyxl.styles.Font(bold=True, color=openpyxl.styles.colors.WHITE)
        sheet.cell(row=1, column=x+1).fill = cabecera
    print(5)
    ## datos color
    for y in range(1,filas):
        for x in range(columnas):
            if y%2==1:
                break
            else:
               sheet.cell(row=y+1, column=x+1).fill = par
    print(6)
    print(os.path.join(dst,name))
    wb.save(filename = os.path.join(dst,name))
    print(7)
    wb.close()
    print(8)
    del wb

def CorreoDespacho():
    Base = """
	SELECT TIPO,COUNT(*) CANTIDAD FROM (
	SELECT NUMERO_RESOLUCION, CASE WHEN CANAL='Digital' THEN 'AYD' ELSE 'GEOXPRESS' END AS TIPO ,
	ROW_NUMBER() over (partition by NUMERO_RESOLUCION,CANAL order by id )DUP
	FROM  corr..robot_despacho
	WHERE FECHA_CARGA=CONVERT(VARCHAR,GETDATE(),103) AND GUIA IS NOT NULL AND NOMBRE_ARCHIVO LIKE '%.PDF%'
	)Q
	WHERE DUP = 1
	GROUP BY TIPO
    """
    Select = """
	SELECT DISTINCT
	CASE WHEN CANAL='Digital' THEN 'AYD' ELSE 'GEOXPRESS' END AS TIPO, 
	NUMERO_RESOLUCION, FECHA_RESOLUCION, COD_RECLAMO, SERVICIO, NOMBRE, ANALISTA, FECHA_DESPACHO, DIRECCION, DISTRITO, PROVINCIA,
	DEPARTAMENTO, NEGOCIO, RESULTADO, CORREO
	FROM robot_despacho WHERE FECHA_CARGA=CONVERT(VARCHAR,GETDATE(),103) AND GUIA IS NOT NULL AND NOMBRE_ARCHIVO LIKE '%.PDF%'
    """
    Lista, fp = DiccionarioSQL(Base, 4), open('Mensaje2.html')
    Subject = "Notificación de Despacho " + dt.today().strftime("%d/%m/%Y")
    Recipient = [a['DESPACHO_ENTREGA_CON PROVEEDOR_CORREO_INFORMATIVO_SIN_EXCEL'] for a in DiccionarioExcelXLSX('Correos.xlsx') if a['DESPACHO_ENTREGA_CON PROVEEDOR_CORREO_INFORMATIVO_SIN_EXCEL']!='']
    print('antes de enviar')
    SMTPLib(Subject,Recipient, fp.read().replace("[REEMPLAZAR]", TablaHTML(Lista)))

    Lista, fp = DiccionarioSQL(Base, 4), open('Mensaje2.html')
    Subject = "Notificación de Despacho " + dt.today().strftime("%d/%m/%Y")
    Recipient = [a['SIN PROVEEDOR_CON_EXCEL'] for a in DiccionarioExcelXLSX('Correos.xlsx') if a['SIN PROVEEDOR_CON_EXCEL']!='']
    attachments = []
    with open(ResumenPrevio_3(DiccionarioSQL(Select, 4), 'Casos Despachados'), 'rb') as f: content = f.read()
    attachments.append(('Casos Despachados.xls', content))

    SMTPLib(Subject,Recipient, fp.read().replace("[REEMPLAZAR]", TablaHTML(Lista)), attachments)
##    SMTPLib(Subject,Recipient, fp.read().replace("[REEMPLAZAR]", TablaHTML(Lista)))

def carga_input():
    global ssh
##  C:\Users\%USERNAME%\AppData\Local\Programs\Python\Python37\lib\site-packages\paramiko\py3compat.py
##  return s.decode(encoding) -> s.decode(encoding, 'ignore') line 147
    def Reconectar(ssh,Ruta):
        ssh.close()
        ssh.connect('172.28.11.228', username="usr_ychavez", password="Password$31",port=22, timeout=100)
        sftp = ssh.open_sftp()
        sftp.chdir(Ruta)
        return sftp
    def get_file(remotePath):
        count=0
        names=[]
        error=[]
        for fileattr in sftp.listdir_attr(remotePath):
            count+=1
            names.append(fileattr.filename)
            if not(stat.S_ISREG(fileattr.st_mode)):error.append({'ERROR':fileattr.filename})
        return count,names,error
    def downLoadFile(ssh,sftp, remotePath, localPath,excepciones=[]):
        global names
        times=3
        count=0
        names=[]
        error=[]
        print("remotePath:",remotePath)
        for fileattr in sftp.listdir_attr(remotePath):
            count+=1
            names.append(fileattr.filename)
            if fileattr.filename in os.listdir(localPath):continue
            if fileattr.filename in excepciones:continue
            if stat.S_ISREG(fileattr.st_mode):
                print(fileattr.filename)
                time=1
                while True:
                    if time==2**times:
                        error.append({'ERROR':fileattr.filename})
                        break
                    try:
                        sftp.get(fileattr.filename, os.path.join(localPath, fileattr.filename))
                        break
                    except Exception as e:
                        if not(ssh.get_transport().is_active()) or sftp.get_channel().closed:
                            try:
                                sftp=Reconectar(ssh,remotePath)
                            except Exception as e:
                                print("Error1:",e)
                                print(f"se espera {time} segundos ....")
                                sleep(time)
                                time*=2
                            continue
                        print("Error2:",e)
                        if 'File not found' in str(e):
                            error.append({'ERROR':fileattr.filename})
                            break
                        print(f"se espera {time} segundos ....")
##                        sleep(time)
                        time*=2
            else:
                print("Nombre de archivo incorrecto:",fileattr.filename)
                error.append({'ERROR':fileattr.filename})
        for x in [y['ERROR'] for y in error]:
            os.remove(os.path.join(localPath,x))
        print("Cantidad de archivos descargados es ",count)
        return error,count,sftp,names
    
    def rename(ssh,sftp,src,dst,times=8):
        time=1
        while True:
            if time==2**times:return True##indica error
            try:
                sftp.rename(src,dst)
                return False
            except Exception as e:
                if not(ssh.get_transport().is_active()) or sftp.get_channel().closed:
                    try:
                        sftp=Reconectar(ssh,dst)
                    except Exception as e:
                        print("Error1:",e)
                        print(f"se espera {time} segundos ....")
                        sleep(time)
                        time*=2
                    continue
                print("Error21:",e)
                if 'File not found' in str(e):return True
                print(f"se espera {time} segundos ....")
                sleep(time)
                time*=2
    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    ssh.connect('172.28.11.228', username="usr_ychavez", password="Password$31",port=22)
    locale.setlocale(locale.LC_ALL, 'es-ES')
    mes= date.today().strftime("%m_%B_%Y")
    mes1=(date.today()-timedelta(days=1 if date.today().isoweekday()!=1 else 3 )).strftime("%m_%B_%Y")
    locale.setlocale(locale.LC_ALL, 'en_US')    
    print(mes)
    sftp = ssh.open_sftp()    
    Ruta = '/SSH/Despacho UE/DESPACHO_PYTHON/INPUT/{1}/{0}'.format((date.today()-timedelta(days=1 if date.today().isoweekday()!=1 else 3 )).strftime("%Y-%m-%d"),mes1)
    Ruta_fin = '/SSH/Despacho UE/DESPACHO_PYTHON/INPUT/{1}/{0}'.format((date.today()).strftime("%Y-%m-%d"),mes)
    RutaSSH(sftp, Ruta)
##    return sftp
    print(1)
    count1,names1,error1=get_file(Ruta)
    print(2)
    RutaSSH(sftp, Ruta_fin)
    print(3)
    count,names,error=get_file(Ruta_fin)
    if error1:
        print("error 1")
##        make_excel(error1,os.getcwd(),'Archivos_nombre_no_correctos.xlsx')
        with open('Archivos_nombre_no_correctos.txt','a') as f:f.write("\r\n".join([x['ERROR'] for x in error1]))
        messagebox.showinfo("Error", f"Hay archivos con el nombre mal escribo en la ruta: {Ruta}")
        return
    elif error:
        print("Error")
##        make_excel(error,os.getcwd(),'Archivos_nombre_no_correctos.xlsx')
        with open('Archivos_nombre_no_correctos.txt','a') as f:f.write("\r\n".join([x['ERROR'] for x in error]))
        messagebox.showinfo("Error", f"Hay archivos con el nombre mal escribo en la ruta: {Ruta_fin}")
        return
    if names1:
        RutaSSH(sftp, Ruta_fin)
        for x in names1:
            if rename(ssh,sftp,f"{Ruta}/{x}",f"{Ruta_fin}/{x}"):error.append({'ERROR':x})
    print(4)
    RutaSSH(sftp, Ruta_fin)
    print(5)
    excepciones=os.listdir(os.path.join(os.getcwd(),'Input',mes,(date.today()).strftime("%Y-%m-%d")))
    error1,count,sftp,names=downLoadFile(ssh,sftp, Ruta_fin,os.path.join(os.getcwd(),'Input',mes,(date.today()).strftime("%Y-%m-%d")),excepciones)
    print(6)
    error+=error1
    print(error,count)
    print("se descargaron los archivos")
    RutaSSH(sftp, Ruta_fin.replace('INPUT','PROCESADOS'))
    for x in names:
        if rename(ssh,sftp,f"{Ruta_fin}/{x}",f"{Ruta_fin.replace('INPUT','PROCESADOS')}/{x}"):error.append({'ERROR':x})
    if error:
##        make_excel(error,os.getcwd(),'Archivos_nombre_no_correctos.xlsx')
        with open('Archivos_nombre_no_correctos.txt','a') as f:f.write("\r\n".join([x['ERROR'] for x in error]))
        messagebox.showinfo("Error", f"Hay archivos con el nombre mal escribo en la ruta: {Ruta_fin}")
        sftp.close()
        ssh.close()
        return
    sftp.close()
    ssh.close()
    messagebox.showinfo("Mensaje", "Terminó")

def Activate():
##    global Base,datos
    print ("ddd")
    Void(os.getcwd()+ "\\Resultado")
    Execute("UPDATE robot_despacho SET NOMBRE_ARCHIVO='ANULADO' WHERE GUIA IS NULL AND FECHA_CARGA!=CONVERT(VARCHAR,GETDATE(),103) AND NOMBRE_ARCHIVO IS NULL", 4)
    #Obtener PDFs
    Pdfs, GrupoPDFs = list(filter(lambda f: f.endswith(('.pdf','.PDF')), [f for f in os.listdir(Ruta_Despacho)])), []
    pb.configure(maximum=len(Pdfs))
    progress.set(1)
    for idy, file in enumerate(Pdfs):
        progress.set(idy+1)
        Cuenta.config(text=str(idy+1)+"/"+str(len(Pdfs)))
        root.wm_attributes("-topmost", True)
        root.update_idletasks()
        root.update()
        try:
            pdf = PdfFileReader(Ruta_Pdfs + "\\" + file)
            pdf._override_encryption = True
            pdf._flatten()
            doc = fitz.open(Ruta_Pdfs + "\\" + file)
            Grupo = [pdf.getPage(a) for a in range(len(range(pdf.getNumPages())))]
            GrupoTexto = [doc.loadPage(a).getText() for a in range(len(range(pdf.getNumPages())))]
            for page in range(pdf.getNumPages()):
                PDF = {}
                PDF['PDF'] = Grupo
                PDF['TEXTO'] = GrupoTexto
                PDF['ESCRITO'] = GrupoTexto[page] #Match
                PDF['GRUPO'] = file
                PDF['PAGINA'] = int(page)
                GrupoPDFs.append(PDF)
        except: print("CASO:", file)
    #print("Termino leer pdfsssssssssssssssssssss")
    #print(GrupoPDFs) 
    #Obtener Excels
    #Repetidos = [str(a['NUMERO_RESOLUCION']) for a in DiccionarioSQL("SELECT DISTINCT NUMERO_RESOLUCION FROM robot_despacho WHERE FECHA_CARGA!=CONVERT(VARCHAR,GETDATE(),103) AND GUIA IS NOT NULL", 4)]
    Excels = list(filter(lambda f: f.endswith(('.xlsx','.XLSX')), [f for f in os.listdir(Ruta_Despacho)]))
    pb.configure(maximum=len(Excels))
    progress.set(1)
    for idy, z in enumerate(Excels):
        try:
            progress.set(idy+1)
            Cuenta.config(text=str(idy+1)+"/"+str(len(Excels)))
            root.wm_attributes("-topmost", True)
            root.update_idletasks()
            root.update()
            print("Excel pre - leído :", z)            
            Base = [Cambio(a) for a in DiccionarioExcelXLSX(Ruta_Despacho + "\\" + z)]
            print("diccionario extraido!")
            if  len([ y for  y in  Base  if 'ERROR' in list(y.keys())])> 0:                            
                print ("Excel error: " , z , " Columna: " ,  [y['ERROR']for y in Base][0])
            else:    
                Base = [Execute("INSERT INTO robot_despacho ({0},FECHA_CARGA,NOMBRE_EXCEL) VALUES ({1},CONVERT(VARCHAR,GETDATE(),103),{2})".format(",".join([str(c) for c in list(b.keys())]), ",".join(["'" + str(b[c]) + "'" if str(c)!='' else 'NULL' for c in list(b)]),"'"+str(z)+"'"), 4) for b in Base if b['NUMERO_RESOLUCION']!='']
                print("Excel leído 1:", z, len(Base))
        except Exception as err:
            print("Excel no leído 2:", z , " Error: ",str(err)); NoExcel.append(z);
            shutil.copy(Ruta_Despacho + "\\{0}".format(z), Ruta_Eliminados + "\\{0}".format(z))

##################################################################################################################

    Excels = list(filter(lambda f: f.endswith(('.xls','.XLS')), [f for f in os.listdir(Ruta_Despacho)]))
    pb.configure(maximum=len(Excels))
    progress.set(1)
    for idy, z in enumerate(Excels):
        try:
            progress.set(idy+1)
            Cuenta.config(text=str(idy+1)+"/"+str(len(Excels)))
            root.wm_attributes("-topmost", True)
            root.update_idletasks()
            root.update()
            #activar para ver que excel no lee o genera error
            print("Excel pre - leído :", z   )
            Base = [Cambio(a) for a in DiccionarioExcelXLS(Ruta_Despacho + "\\" + z)]
            if  len([ y for  y in  Base  if 'ERROR' in list(y.keys())])> 0:            
                print ("Excel error: " , z , " Columna: " ,  [y['ERROR']for y in Base][0])
            else:    
                Base = [Execute("INSERT INTO robot_despacho ({0},FECHA_CARGA,NOMBRE_EXCEL) VALUES ({1},CONVERT(VARCHAR,GETDATE(),103),{2})".format(",".join([str(c) for c in list(b.keys())]), ",".join(["'" + str(b[c]) + "'" if str(c)!='' else 'NULL' for c in list(b)]),"'"+str(z)+"'"), 4) for b in Base if b['NUMERO_RESOLUCION']!=''  ]            
                print("Excel leído 1:", z, len(Base))
        except Exception as err:
            print("Excel no leído 2:", z , " Error: ",str(err)); NoExcel.append(z);
            shutil.copy(Ruta_Despacho + "\\{0}".format(z), Ruta_Eliminados + "\\{0}".format(z))

################################## Duplicado ##################################3
    
    #Validaciones
    #trae todo lo que se subio hoy
    Base = DiccionarioSQL("SELECT ID,NUMERO_RESOLUCION,ANALISTA,NOMBRE_EXCEL,NOMBRE,FECHA_RECLAMO FROM robot_despacho WHERE GUIA IS NULL AND NOMBRE_ARCHIVO IS NULL", 4)
    #resoluciones unicas 
    #messagebox.showinfo("Mensaje", "Terminó")
    #return ''
    #Resoluciones = list(dict.fromkeys([x['NUMERO_RESOLUCION'] for x in Base if x['NUMERO_RESOLUCION']!=''])) #Duplicados
    #for x in Resoluciones:
    #    SubResoluciones = list(filter(lambda a: a['NUMERO_RESOLUCION']==x, Base))
    #    if SubResoluciones != []:
    #        Fechas = list(dict.fromkeys([z['FECHA_RECLAMO'] for z in SubResoluciones]))
    #        for idy, y in enumerate(['ANALISTA','NOMBRE_EXCEL','NOMBRE']):
    #            #print(len(Fechas))
    #            Grupo = list(dict.fromkeys([z[y] for z in SubResoluciones]))
    #            #marca como duplicado si se subio una misma resolucion pero con diferente analista,nombre_excel , nombre y que sea una sola fecha
    #            if len(Grupo)>1 and len(Fechas)==1:
    #                #actualiza como duplicado toda la base y sale de bucle
    #                Execute("UPDATE robot_despacho SET NOMBRE_ARCHIVO='DUPLICADO' WHERE NUMERO_RESOLUCION='{0}'".format(x), 4) ; break;
    #                #·Solo despachar un duplicado
    #                ID = "SELECT ID FROM robot_despacho WHERE NOMBRE_ARCHIVO='DUPLICADO' AND NUMERO_RESOLUCION='{0}' AND FECHA_RECLAMO='{1}'".format(x, str(Fechas[0]))
    #                ID = DiccionarioSQL(ID, 4)
    #                if len(ID) == 0: continue
    #                print('Recuperando Duplicado:', ID)
    #                Execute("UPDATE robot_despacho SET NOMBRE_ARCHIVO=NULL,guia = null  WHERE ID='{0}'".format(str(ID[0]['ID'])),4)
                
    Execute("EXEC web_visor..SP_SA_VALIDACIONES_DESPACHO")
    Base, Todo = DiccionarioSQL(Robot_Despacho, 4), []
##    data=('BRC1118568', 'BRC1119185', 'BRC1119324', 'BRC1118308', 'BRC1118975', 'BRC1118959', 'BRC1117827', 'BRC1118444', 'BRC1117842', 'BRC1119294', 'BRC1118520', 'BRC1118769', 'BRC1118342', 'BRC1118091', 'BRC1117891', 'BRC1118803', 'BRC1118254', 'BRC1118648', 'BRC1118825', 'BRC1118293', 'BRC1118408', 'BRC1118385', 'BRC1118317', 'BRC1118056', 'BRC1118075', 'BRC1117942', 'BRC1118624', 'BRC1119265', 'BRC1118742', 'BRC1118766', 'BRC1118465', 'BRC1119086', 'BRC1118684', 'BRC1118237', 'BRC1118345', 'BRC1118755', 'BRC1117753', 'BRC1119178', 'BRS0278380', 'BRC1118149', 'BRTF001612', 'BRC1118676', 'BRS0278478', 'BRC1118050', 'BRC1117985', 'BRC1117609', 'BRC1118661', 'BRC1118011', 'BRC1118121', 'BRC1118309', 'BRC1118776')
##    datos=[x for x in Base if x['COD_RECLAMO'] in data]
    
    Sql = "SELECT NUMERO_RESOLUCION FROM CORR..robot_despacho WHERE FECHA_CARGA = CONVERT(VARCHAR,GETDATE(),103) AND NOMBRE_ARCHIVO IS NOT NULL"
    Base_despachados = [a['NUMERO_RESOLUCION']  for a in DiccionarioSQL(Sql, 4) if a['NUMERO_RESOLUCION']!='']     
    #Guias y PDFs
    def Formato(stra):
        if "/" in stra and len(stra)==10:
            stra = stra.split("/")
            return stra[2]+stra[1].zfill(2)+stra[0].zfill(2)
        else: return stra

    for x in ['Fundado', 'Infundado']:
        Guia = str(str(GuiaMax() + (1 if x.upper()=='INFUNDADO' else 2))).zfill(3)
        Lista = [a for a in Base if a['RESULTADO'].upper()==x.upper()]
        pb.configure(maximum=len(Lista))
        progress.set(1)
        for idy, y in enumerate([a for a in Base if a['RESULTADO'].upper()==x.upper()]):
            progress.set(idy+1)
            Cuenta.config(text=str(idy+1)+"/"+str(len(Lista)))
            root.wm_attributes("-topmost", True)
            root.update_idletasks()
            root.update()
            Pdf = [a for a in GrupoPDFs if y['NUMERO_RESOLUCION'].replace(" ", "") in a['ESCRITO'].replace(" ", "")]
            if len(Pdf)>0:
                Pdf = Pdf[0]
                Pdf['CARTA'] = []
                Codigos = [a['NUMERO_RESOLUCION'].replace(" ", "") for a in Base if str(a['NUMERO_RESOLUCION']).replace(" ", "")!=str(y['NUMERO_RESOLUCION']).replace(" ", "")]
                for idz, z in enumerate(Pdf['PDF'][Pdf['PAGINA']:]):
                    Pdf['CARTA'].append(z)
                    if idz>3:
                        #verificar si esto no deberia salir
                        print('CARTA +3 HOJAS NO EXISTE:', y['NUMERO_RESOLUCION'],  ' pdf: ', Pdf['GRUPO'])
                        Pdf['CARTA'] = [Pdf['PDF'][Pdf['PAGINA']]]; break;
                    if len(Pdf['PDF'])==Pdf['PAGINA']+idz+1: break;
                    Siguiente = str(Pdf['TEXTO'][Pdf['PAGINA']+idz+1]).replace(" ", "")
                    if len([a for a in Codigos if a in Siguiente])>0: break;
                    if len([a for a in Base_despachados  if a in Siguiente])>0: break;                    
                y['PDF'] = Pdf['CARTA']
                y['GUIA'] = Guia + "-" + str(idy).zfill(3) + "-" + str(datetime.now().year)
                y['NOMBRE_PDF'] = y['NUMERO_RESOLUCION'] + ".pdf"
            else: #PDFs no encontrados
                y['PDF'], y['GUIA'], y['NOMBRE_PDF'] = [], "", ""
                Execute("UPDATE robot_despacho SET NOMBRE_ARCHIVO='NO EXISTE PDF' WHERE ID={0}".format(y['ID']), 4)
            Todo.append(y)
    #Agrupar TODO

    if len(Todo)>0:        
        for x in ['Fisico', 'Digital']:
            Errores = ['DUPLICADO','ANULADO','NO EXISTE PDF']
            Despacho = [a for a in Todo if a['GUIA']!='' and a['NOMBRE_ARCHIVO'] not in Errores and str(a['CANAL']).upper()==x.upper() and "/" not in a['NUMERO_RESOLUCION']]
            if len(Despacho)>0:
                os.mkdir(os.getcwd()+ "\\Resultado\\" + x)
                [CreaPDF(a) for a in Despacho]
                Consolidado(AgruparTODO(Despacho))
        Execute("UPDATE robot_despacho SET NOMBRE_ARCHIVO='FORMATO' WHERE NOMBRE_ARCHIVO IS NULL AND FECHA_CARGA=CONVERT(VARCHAR,GETDATE(),103)", 4)
        print('Listo!')
    else: print("No hay data nueva")

    messagebox.showinfo("Mensaje", "Terminó")

#Validaciones
def AgruparTODO(Todo):
    Resultado, Resoluciones = [], list(dict.fromkeys([x['NUMERO_RESOLUCION'] for x in Todo if x['NUMERO_RESOLUCION']!='']))
    Linea = ['COD_RECLAMO','FECHA_RECLAMO','NOMBRE']
    for idx, x in enumerate(Resoluciones):
        SubResoluciones = [z for z in Todo if z['NUMERO_RESOLUCION'] == x]
        if SubResoluciones != []:
            Nuevo = {key: SubResoluciones[0][key] for key in SubResoluciones[0].keys() if key not in Linea} #No diferentes
            for idy, y in enumerate(Linea): #Diferentes
                 Nuevo[y] = ",".join(list(dict.fromkeys([z[y] for z in SubResoluciones])))
            Resultado.append(Nuevo)         
    return Resultado

def Actualizar():
    fname = askopenfilename(filetypes=(("Excel files", "*.xlsx"),("All files", "*.*")))
    if fname:
        def DiccionarioExcel(Excel, Filass=1):
            Writer = openpyxl.load_workbook(Excel)
            Sheet = Writer.active
            Header, Listado = list(n.value for n in Sheet[Filass]), []
            pb.configure(maximum=int(Sheet.max_row)); Max = Sheet.max_row;
            progress.set(1)
            for idx in range(Max):
                Lista = {}
                progress.set(idx+1)
                Cuenta.config(text=str(idx+1)+"/"+str(int(Max)))
                root.wm_attributes("-topmost", True)
                root.update_idletasks()
                root.update()
                for idy, y in enumerate(Header):
                    if y == None: continue
                    value = Sheet.cell(row=idx+1+Filass, column=idy+1).value
                    value = value if value != None and value != '' else ''
                    if 'a. m.' in str(value) or 'p. m.' in str(value) or 'AM' in str(value) or 'PM' in str(value) and 'FECHA' in y.upper():
                        value = str(value).replace('p. m.', 'pm').replace('a. m.', 'am').replace('AM', 'am').replace('PM', 'pm')
                        value = str(value).split(' ')
                        if len(value)==4: value = value[0] + " " + value[1] + ":" + value[2].zfill(2) + " " + value[3]
                        elif len(value)==3:
                            if len(value[1].split(":"))==2: val = value[2].split(':'); value = value[0] + " " + value[1] + ":" + val[0].zfill(2) + " " + val[1]
                            else: value = value[0] + " " + value[1] + " " + value[2]
                        else: value = ''
                    if 'p. m.' in str(value): value = str(value).replace('p. m.', 'pm')
                    if isinstance(value, datetime): value = value.strftime("%Y%m%d %H:%M:%S")
                    elif ('am' in str(value) or 'pm' in str(value)) and ':' in str(value) and 'FECHA' in y.upper(): value = datetime.strptime(str(value), '%d/%m/%Y %H:%M:%S %p').strftime("%Y%m%d %H:%M:%S");
                    elif len(str(value))==10 and "-" in str(value):
                        value = str(value).split("-")
                        if len(value)==3: value = value[0] + value[1].zfill(2) + value[2].zfill(2)
                        else: value = ''
                    else: value = str(value)
                    Lista[y.upper()] = str(value).replace("'", "")
                Listado.append(Lista)
            return Listado
        A = DiccionarioExcel(fname, 2)
        print('Inicio!')
        pb.configure(maximum=len(A))
        progress.set(1)
        Grupo = [a['RESOLUCION'] + "_" + a['PRODUCTO'] for a in DiccionarioSQL("SELECT RESOLUCION, PRODUCTO FROM [robot_reporte_distribucion]", 4)]
        for ida, a in enumerate(A):
            if a["RESOLUCIÓN"] + "_" + a["PRODUCTO"] in Grupo: continue;
            if a["RESOLUCIÓN"]=='': break
            try:
                progress.set(ida+1)
                Cuenta.config(text=str(ida+1)+"/"+str(len(A)))
                root.wm_attributes("-topmost", True)
                root.update_idletasks()
                root.update()
                Nuevo = {}
                Nuevo['CODIGO_GEO'] = a['CODIGO GEO']
                Nuevo['GUIA'] = a['GUIA']
                Nuevo['RESOLUCION'] = a['RESOLUCIÓN']
                Nuevo['CICLO'] = a['CICLO']
                Nuevo['DETALLE_ESTADO'] = a['DETALLE ESTADO']
                Nuevo['FECHA_ENTREGA'] = a['FECHA ENTREGA']
                Nuevo['TIPO_ZONA'] = a['TIPO ZONA']
                Nuevo['PRODUCTO'] = a['PRODUCTO']
                Nuevo['FECHA_DIGITALIZACION'] = a['FECHA DE DIGITALIZACIÓN']
                Nuevo['PROCEDENCIA'] = a['PROCEDENCIA']
                Nuevo['TELEFONO'] = a['TELÉFONO']
                Nuevo['CODIGO_RECLAMO'] = a['RECLAMO']
                Execute("INSERT INTO [robot_reporte_distribucion] ({0},FECHA_CARGA) VALUES({1},GETDATE())".format(",".join(list(Nuevo.keys())), ",".join([ "'" + c + "'" if c!='' else "NULL" for c in list(Nuevo.values())])), 4)
            except: print("INSERT INTO [robot_reporte_distribucion] ({0}) VALUES({1})".format(",".join(list(Nuevo.keys())), ",".join([ "'" + c + "'" if c!='' else "NULL" for c in list(Nuevo.values())])))
        messagebox.showinfo("Mensaje", "Terminó")

def RutaSSH(sftp, Ruta):
    Ruteando = [a for a in Ruta.split("/") if a!='']
    for idx, x in enumerate(Ruteando):
        try: sftp.chdir("".join(["/{0}".format(str(b)) for b in Ruteando[:idx+1]]))
        except IOError:
            sftp.mkdir("".join(["/{0}".format(str(b)) for b in Ruteando[:idx+1]]))
            sftp.chdir("".join(["/{0}".format(str(b)) for b in Ruteando[:idx+1]]))

def SubirSSH():
    def Reconectar(ssh,Ruta):
        ssh.close()
        ssh.connect('172.28.11.228', username="usr_ychavez", password="Password$31",port=22, timeout=100)
        sftp = ssh.open_sftp()
        sftp.chdir(Ruta)
        return sftp
    def put(ssh,sftp,src,dst,times=8):
        time=1
        while True:
            if time==2**times:break
            try:
                sftp.put(src,dst)
                break
            except Exception as e:
                if not(ssh.get_transport().is_active()) or sftp.get_channel().closed:
                    try:
                        sftp=Reconectar(ssh,dst)
                    except Exception as e:
                        print("Error1:",e)
                        print(f"se espera {time} segundos ....")
                        sleep(time)
                        time*=2
                    continue
                print("Error2:",e)
                print(f"se espera {time} segundos ....")
                sleep(time)
                time*=2
                
    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    User = DiccionarioSQL("SELECT APLICATIVO, USUARIO, CLAVE FROM robot_accesibilidad WHERE Aplicativo = 'SSH'", 4)[0]
    print(User)
    #ssh.connect('172.28.11.228', username=User['USUARIO'], password=User['CLAVE'])
##    ssh.connect('172.28.11.228', username="lurcuhua", password="Brasil54+",port=22, timeout=100)
    ssh.connect('172.28.11.228', username="usr_ychavez", password="Password$28",port=22, timeout=100)
##    ssh.connect('172.28.11.228', username="ssh_jsantiagoa", password="Jhon%123456789023")
    sftp = ssh.open_sftp()
    locale.setlocale(locale.LC_ALL, 'es-ES')
    mes= date.today().strftime("%m")+"."+date.today().strftime("%B").upper()
    locale.setlocale(locale.LC_ALL, 'en_US')    
    print(mes)

    for x in ['Fisico', 'Digital']:
        
        if not os.path.isdir(os.getcwd() + "//Resultado//{0}".format(x)): continue
        Ruta = '/SSH/Despacho UE/DESPACHO_ROBOT/{0}/{1}/{2}'.format(mes,x,date.today().strftime("%d.%m.%y"))
        RutaSSH(sftp, Ruta)
        Excel = '{0} {1}.xls'.format(x, date.today().strftime("%d.%m.%Y"))
        put(ssh,sftp,os.getcwd() + "\\Resultado\\{0}".format(Excel),"{0}/{1}".format(Ruta, Excel))
##        sftp.put(os.getcwd() + "\\Resultado\\{0}".format(Excel), "{0}/{1}".format(Ruta, Excel)) #Mover Excel
        Lista = os.listdir(os.getcwd() + "/Resultado/{0}".format(x))
        pb.configure(maximum=len(Lista))
        progress.set(1)
        Carpeta = '{0} {1}'.format(x, date.today().strftime("%d.%m.%Y"))
        Carpeta = '{0}'.format(x)
        
        for idy, y in enumerate(Lista):
            progress.set(idy+1)
            Cuenta.config(text=str(idy+1)+"/"+str(len(Lista)))
            root.wm_attributes("-topmost", True)
            root.update_idletasks()
            root.update()
            if idy==0: RutaSSH(sftp, Ruta + "/" + Carpeta);
            print (Ruta)
            print (Carpeta)
            print (y)
            put(ssh,sftp,os.getcwd() + "/Resultado/{0}/{1}".format(x, y), "{0}/{1}/{2}".format(Ruta, Carpeta, y))
##            sftp.put(os.getcwd() + "/Resultado/{0}/{1}".format(x, y), "{0}/{1}/{2}".format(Ruta, Carpeta, y)) #Mover Archivos
    sftp.close()
    ssh.close()
    messagebox.showinfo("Mensaje", "Terminó")

def AnularValores():
    fname = askopenfilename(filetypes=(("Excel files", "*.xlsx"),("All files", "*.*")))
    if fname:
        Lista = [a for a in DiccionarioExcelXLSX(fname) if a['NUMERO_RESOLUCION']!='']
        pb.configure(maximum=len(Lista))
        progress.set(1)
        for idx, x in enumerate(Lista):
            progress.set(idx+1)
            Cuenta.config(text=str(idx+1)+"/"+str(len(Lista)))
            root.wm_attributes("-topmost", True)
            root.update_idletasks()
            root.update()
            print(x['NUMERO_RESOLUCION'])
            Execute("UPDATE robot_despacho SET NOMBRE_ARCHIVO='ANULADO',GUIA=NULL WHERE NUMERO_RESOLUCION='{0}'".format(x['NUMERO_RESOLUCION']), 4)
            Execute("UPDATE robot_cartas_movil_primera SET DIRECCION=NULL, FECHA_GENERADA=NULL, SPEECH=NULL, NUMERO_RESOLUCION=NULL WHERE NUMERO_RESOLUCION='{0}'".format(x['NUMERO_RESOLUCION']))
            Execute("UPDATE robot_cartas_movil_segunda SET DIRECCION=NULL, FECHA_GENERADA=NULL, SPEECH=NULL, NUMERO_RESOLUCION=NULL WHERE NUMERO_RESOLUCION='{0}'".format(x['NUMERO_RESOLUCION']))
            Execute("UPDATE robot_cartas_fija_primera SET DIRECCION=NULL, FECHA_GENERADA=NULL, SPEECH=NULL, NUMERO_RESOLUCION=NULL WHERE NUMERO_RESOLUCION='{0}'".format(x['NUMERO_RESOLUCION']))
        messagebox.showinfo("Mensaje", "Terminó")

def LecturaPDF():
    fname = askopenfilename(filetypes=(("PDF files", "*.pdf"),("All files", "*.*")))
    if not fname: return;
    Ruta = os.path.split(fname)[0].replace("/", "\\")
    Lista = list(filter(lambda f: f.endswith(('.pdf','.PDF')), [f for f in os.listdir(Ruta)]))
    pb.configure(maximum=len(Lista))
    progress.set(1)
    for idx, file in enumerate(Lista):
        progress.set(idx+1)
        Cuenta.config(text=str(idx+1)+"/"+str(len(Lista)))
        root.wm_attributes("-topmost", True)
        root.update_idletasks()
        root.update()
        doc = fitz.open(Ruta + "\\" + file)
        Telefono = doc.loadPage(0).getText().split('Cód.pago: ')[1].split('\n')[0]
        Base = DiccionarioSQL("SELECT TOP 1 COD_RECLAMO, ID FROM robot_Cartas_fija_primera WHERE SERVICIO='{0}'".format(Telefono))
        if len(Base)==0: print('NO HECHO:', file); continue
        shutil.copy(Ruta + "\\{0}".format(file), Ruta_descargas + "\\{0}.pdf".format(Base[0]['COD_RECLAMO']))

if __name__ == '__main__':
    NoExcel = []
    Robot_Despacho = """
        SELECT ID,NEGOCIO,NUMERO_RESOLUCION,FECHA_RESOLUCION,COD_RECLAMO,SERVICIO,NOMBRE,
        ANALISTA,FECHA_DESPACHO,DIRECCION,DISTRITO,PROVINCIA,DEPARTAMENTO,ANEXO_INSCRIPCION,CLIENTE,SEGMENTO,FECHA_RECLAMO,
        MONTO_RECLAMO,TIPO_ENVIO,RESULTADO,CORREO,CANAL,INSTANCIA,CODIGO_RECLAMO_1RA,FECHA_CARGA,GUIA,NOMBRE_ARCHIVO
        FROM robot_despacho
        WHERE GUIA IS NULL AND NOMBRE_ARCHIVO IS NULL
        ORDER BY RESULTADO ASC
    """
    Robot_No_Despachado = "SELECT NEGOCIO,ANALISTA,NOMBRE_EXCEL,NUMERO_RESOLUCION,COD_RECLAMO,FECHA_RECLAMO,NOMBRE_ARCHIVO AS OBSERVACION FROM ROBOT_DESPACHO WHERE GUIA IS NULL AND FECHA_CARGA=CONVERT(VARCHAR,GETDATE(),103)"

    locale.setlocale(locale.LC_ALL, 'es-ES')
    mes= date.today().strftime("%m")+"_"+date.today().strftime("%B").upper()+"_"+date.today().strftime("%Y")
##    Ruta_Despacho = "\\\\10.4.40.191\\Informacion Primera y Segunda Instancia Soluciones\\DESPACHO_ROBOT\\Input\\{1}\\{0}".format(date.today().strftime("%Y-%m-%d"),mes)
    Ruta_Despacho = os.getcwd()+"\\Input\\{1}\\{0}".format(date.today().strftime("%Y-%m-%d"),mes)
##def prueba():
    if not(os.path.isdir(os.getcwd()+"\\Input\\{0}".format(mes))):os.mkdir(os.getcwd()+"\\Input\\{0}".format(mes))
    if not(os.path.isdir(Ruta_Despacho)):os.mkdir(Ruta_Despacho)
    #Ruta_Despacho ="D:\\2020-08-26"

    locale.setlocale(locale.LC_ALL, 'en_US')
    
    print(Ruta_Despacho)


    Ruta_Pdfs = Ruta_Despacho
    Ruta_descargas = os.getcwd()+ "\\Descargas"
    Ruta_Eliminados = os.getcwd()+ "\\Eliminados"
    Change = {"NUMERO_RESOLUCION":"NUMERO_RESOLUCIONssssssssssssssss","FECHA_RESOLUCION":"FECHA_RESOLUCION","CODIGO_RECLAMO":"COD_RECLAMO","SERVICIO":"SERVICIO","NOMBRE DEL RECLAMANTE":"NOMBRE","ANALISTA":"ANALISTA","DIRECCION":"DIRECCION","DISTRITO":"DISTRITO","PROVINCIA":"PROVINCIA","DEPARTAMENTO":"DEPARTAMENTO","ANEXO/INSCRIPCION":"ANEXO_INSCRIPCION","FECHA_RECLAMO":"FECHA_RECLAMO","MONTO_RECLAMO":"MONTO_RECLAMO","TIPO_ENVIO":"TIPO_ENVIO","RESULTADO":"RESULTADO","CORREO":"CORREO","CANAL":"CANAL","INSTANCIA":"INSTANCIA","NEGOCIO":"NEGOCIO","CODIGODERECLAMO1RAINSTANCIA":"CODIGO_RECLAMO_1RA"}
    Variables = ['GUIA','NUMERO_RESOLUCION','FECHA_RESOLUCION','COD_RECLAMO','SERVICIO','NOMBRE','ANALISTA','FECHA_DESPACHO','DIRECCION','DISTRITO','PROVINCIA','DEPARTAMENTO','ANEXO_INSCRIPCION','SEGMENTO','FECHA_RECLAMO','CANAL','INSTANCIA','CORREO']
    Diccionario_columnas = DiccionarioSQL("select  Columna_variable,Columna_correcta from sa_diccionario_columnas  ", 4)
    
    #Tkinter
    warnings.filterwarnings("ignore")
    root = tkinter.Tk()
    root.geometry('1220x90+'+str(0)+'+'+str(win32api.GetSystemMetrics(1)-180))
    root.title('Robot')
    root.wm_attributes("-topmost", True)
    #root.overrideredirect(1)
    root.resizable(False, False)

    importar = tkinter.Button(root, text='Iniciar', command=Activate, height = 2, width = 15,bg = "blue", fg = "white")
    importar.place(x=10, y=10)

    Recibos = tkinter.Button(root, text='Recibos', command=Recibos, height = 2, width = 15,bg = "green", fg = "white")
    Recibos.place(x=130, y=10)

    ActRespuesta = tkinter.Button(root, text='Actualizar', command=Actualizar, height = 2, width = 15,bg = "orange", fg = "white")
    ActRespuesta.place(x=250, y=10)

    Faltan = tkinter.Button(root, text='Casos no Despachados', command=Correo_No_Despacho, height = 2, width = 15,bg = "brown", fg = "white")
    Faltan.place(x=370, y=10)

    Despacha = tkinter.Button(root, text='Correo Despacho', command=CorreoDespacho, height = 2, width = 15,bg = "gray", fg = "white")
    Despacha.place(x=490, y=10)

    SSH = tkinter.Button(root, text='Subir SSH', command=SubirSSH, height = 2, width = 15,bg = "black", fg = "white")
    SSH.place(x=610, y=10)

    Anular = tkinter.Button(root, text='Anular', command=AnularValores, height = 2, width = 15,bg = "turquoise", fg = "white")
    Anular.place(x=730, y=10)

    Arreglar = tkinter.Button(root, text='Arreglar Recibos', command=LecturaPDF, height = 2, width = 15,bg = "purple", fg = "white")
    Arreglar.place(x=850, y=10)

    descarga = tkinter.Button(root, text='Descarga ssh', command=carga_input, height = 2, width = 15,bg = "yellow", fg = "red")
    descarga.place(x=970, y=10)


    cerrar = tkinter.Button(root, text='Cerrar', command=Cerrar, height = 2, width = 15,bg = "red", fg = "white")
    cerrar.place(x=1090, y=10)

    Cuenta = tkinter.Label(root, text="??? / ???")
    Cuenta.place(x=160, y=60)

    progress = tkinter.IntVar()
    pb = ttk.Progressbar(root, orient="horizontal", length=120, mode="determinate", variable=progress)
    pb.place(x=10, y=60)
    root.mainloop()
